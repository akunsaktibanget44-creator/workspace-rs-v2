from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Depends, Response, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import io
import logging
import uuid
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, BeforeValidator
from typing import List, Optional, Annotated, Any
from datetime import datetime, timezone, date, timedelta
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from auth import build_auth_router, seed_admin, make_auth_dependencies, get_current_user_optional
from monitoring import build_monitoring_router
from strategy import build_strategy_router
from pdf_export import build_raport_pdf

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Workspace Ruang Sanad API")
api_router = APIRouter(prefix="/api")

get_current_user, require_spv = make_auth_dependencies(db)


# ============ SCOPE HELPERS ============
async def user_scope(user: dict) -> dict:
    """Return effective scope for a user: {is_spv, user_id, anggota_id, divisi_id}"""
    is_spv = user.get("role") == "spv"
    anggota_id = user.get("anggota_id")
    divisi_id = None
    if anggota_id:
        a = await db.anggota.find_one({"id": anggota_id}, {"_id": 0})
        if a:
            divisi_id = a.get("divisi_id")
    return {"is_spv": is_spv, "user_id": user["user_id"], "anggota_id": anggota_id, "divisi_id": divisi_id}


async def _assert_task_access(task_id: str, user: dict) -> dict:
    """Load a task and verify user is SPV, penerima, pemberi, or task belongs to their divisi (rutin only)."""
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(404, "Task not found")
    scope = await user_scope(user)
    if scope["is_spv"]:
        return task
    my_ang = scope["anggota_id"]
    is_rutin = task.get("kategori") in ("HARIAN", "MINGGUAN", "BULANAN")
    if my_ang and (task.get("penerima_tugas_id") == my_ang or task.get("pemberi_id") == my_ang):
        return task
    if is_rutin and scope["divisi_id"] and task.get("divisi_id") == scope["divisi_id"]:
        return task
    raise HTTPException(403, "Tugas ini bukan bagian dari workspace Anda.")


async def _assert_bulk_task_access(ids: List[str], user: dict) -> List[dict]:
    scope = await user_scope(user)
    if not ids:
        return []
    rows = await db.tasks.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "divisi_id": 1}).to_list(len(ids))
    if not scope["is_spv"]:
        if not scope["divisi_id"]:
            raise HTTPException(403, "Akun belum terhubung ke tim.")
        for r in rows:
            if r.get("divisi_id") != scope["divisi_id"]:
                raise HTTPException(403, "Ada tugas di luar tim Anda dalam seleksi.")
    return rows


PUBLIC_PATHS = {"/api", "/api/"}
AUTH_PUBLIC_PREFIXES = ("/api/auth/register", "/api/auth/login", "/api/auth/google", "/api/auth/me", "/api/auth/logout")


@app.middleware("http")
async def require_auth_middleware(request: Request, call_next):
    path = request.url.path
    if not path.startswith("/api/") and path != "/api":
        return await call_next(request)
    if path in PUBLIC_PATHS or path.startswith(AUTH_PUBLIC_PREFIXES):
        return await call_next(request)
    # All other /api/* require an active session
    user = await get_current_user_optional(request, db)
    if not user:
        return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    if user.get("status") != "approved":
        return JSONResponse({"detail": f"Akun {user.get('status', 'pending')} — belum bisa akses."}, status_code=403)
    return await call_next(request)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ============ TASK MODELS ============
class TaskBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    nama: str
    kategori: str = "HARIAN"  # HARIAN | MINGGUAN | BULANAN | PROJECT
    frekuensi: str = "RUTIN"  # RUTIN | TIDAK RUTIN | SEKALI
    status: str = "BELUM_MULAI"  # BELUM_MULAI | DALAM_PROSES | SELESAI | TERKENDALA
    pemberi_tugas: Optional[str] = ""
    penerima_tugas: Optional[str] = ""
    tanggal_mulai: Optional[str] = None  # ISO date
    deadline: Optional[str] = None  # ISO date
    durasi: Optional[str] = ""
    catatan_tim: Optional[str] = ""
    catatan_spv: Optional[str] = ""
    link_output: Optional[str] = ""
    divisi: Optional[str] = "Umum"
    divisi_id: Optional[str] = None
    urutan: int = 0
    list_id: Optional[str] = None
    label_ids: List[str] = Field(default_factory=list)
    kategori_id: Optional[str] = None
    penerima_tugas_id: Optional[str] = None
    pemberi_id: Optional[str] = None  # anggota_id yg mendelegasikan (auto-set on create for anggota)
    brief_link: Optional[str] = ""  # URL brief eksternal (opsional)
    hasil_link: Optional[str] = ""  # URL hasil dari penerima
    hasil_catatan: Optional[str] = ""  # Catatan hasil dari penerima
    revisi_catatan: Optional[str] = ""  # Catatan revisi dari pemberi/SPV
    revisi_at: Optional[str] = None
    revisi_count: int = 0
    moved_at: Optional[str] = None


class Task(TaskBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)
    archived: bool = False
    archived_at: Optional[str] = None


class TaskUpdate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    nama: Optional[str] = None
    kategori: Optional[str] = None
    frekuensi: Optional[str] = None
    status: Optional[str] = None
    pemberi_tugas: Optional[str] = None
    penerima_tugas: Optional[str] = None
    tanggal_mulai: Optional[str] = None
    deadline: Optional[str] = None
    durasi: Optional[str] = None
    catatan_tim: Optional[str] = None
    catatan_spv: Optional[str] = None
    link_output: Optional[str] = None
    divisi: Optional[str] = None
    divisi_id: Optional[str] = None
    urutan: Optional[int] = None
    list_id: Optional[str] = None
    label_ids: Optional[List[str]] = None
    archived: Optional[bool] = None
    kategori_id: Optional[str] = None
    penerima_tugas_id: Optional[str] = None
    pemberi_id: Optional[str] = None
    brief_link: Optional[str] = None
    hasil_link: Optional[str] = None
    hasil_catatan: Optional[str] = None
    revisi_catatan: Optional[str] = None
    revisi_at: Optional[str] = None
    revisi_count: Optional[int] = None
    moved_at: Optional[str] = None


class MoveTaskPayload(BaseModel):
    divisi_id: Optional[str] = None
    list_id: Optional[str] = None
    urutan: Optional[int] = None
    penerima_tugas_id: Optional[str] = None  # pindah/delegasi ke orang lain (boleh lintas divisi)


# ============ DIVISI (TIM) MODEL ============
class Divisi(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama: str
    warna: str = "#10b981"
    urutan: int = 0
    created_at: str = Field(default_factory=now_iso)


class DivisiCreate(BaseModel):
    nama: str
    warna: Optional[str] = "#10b981"


# ============ KATEGORI (USER-MANAGED) MODEL ============
class Kategori(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama: str
    warna: str = "#8b5cf6"
    urutan: int = 0
    created_at: str = Field(default_factory=now_iso)


class KategoriCreate(BaseModel):
    nama: str
    warna: Optional[str] = "#8b5cf6"


# ============ ANGGOTA (MEMBER PER DIVISI) MODEL ============
class Anggota(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    divisi_id: str
    nama: str
    warna: str = "#0ea5e9"
    urutan: int = 0
    created_at: str = Field(default_factory=now_iso)


class AnggotaCreate(BaseModel):
    divisi_id: str
    nama: str
    warna: Optional[str] = "#0ea5e9"


class BulkIdsRequest(BaseModel):
    ids: List[str]


class BulkMoveRequest(BaseModel):
    ids: List[str]
    divisi_id: Optional[str] = None
    list_id: Optional[str] = None


class ReorderRequest(BaseModel):
    task_ids: List[str]


# ============ TASK LIST & LABEL MODELS ============
class TaskList(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama: str
    warna: str = "#10b981"
    urutan: int = 0
    is_done: bool = False
    divisi_id: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)


class TaskListCreate(BaseModel):
    nama: str
    warna: Optional[str] = "#10b981"
    is_done: Optional[bool] = False
    urutan: Optional[int] = None
    divisi_id: Optional[str] = None


class TaskLabel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama: str
    warna: str = "#f59e0b"
    created_at: str = Field(default_factory=now_iso)


class TaskLabelCreate(BaseModel):
    nama: str
    warna: Optional[str] = "#f59e0b"


class TodoEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    task_id: str
    period: str  # YYYY-MM-DD (harian) | YYYY-Www (mingguan) | YYYY-MM (bulanan)
    checked: bool = True
    created_at: str = Field(default_factory=now_iso)


class TodoEntryCreate(BaseModel):
    task_id: str
    period: str
    checked: bool = True


# ============ AMALIYAH MODELS ============
class AmaliyahItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nama: str
    target_metrik: str = "1x/hari"
    keterangan: Optional[str] = ""
    urutan: int = 0
    aktif: bool = True
    created_at: str = Field(default_factory=now_iso)


class AmaliyahItemCreate(BaseModel):
    nama: str
    target_metrik: str = "1x/hari"
    keterangan: Optional[str] = ""


class AmaliyahEntry(BaseModel):
    """Single check-in for an amaliyah on a specific date, scoped to a user."""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_id: str
    tanggal: str  # YYYY-MM-DD
    checked: bool = True
    catatan: Optional[str] = ""
    user_id: Optional[str] = None
    created_at: str = Field(default_factory=now_iso)


class AmaliyahEntryCreate(BaseModel):
    item_id: str
    tanggal: str
    checked: bool = True
    catatan: Optional[str] = ""


# ============ RAPORT MODELS ============
class RaportNote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "singleton"
    catatan_spv: str = ""
    rekomendasi: str = "NETRAL"  # REWARD | EVALUASI | NETRAL
    updated_at: str = Field(default_factory=now_iso)


class RaportNoteUpdate(BaseModel):
    catatan_spv: str = ""
    rekomendasi: str = "NETRAL"


# ============ TASK ROUTES ============
@api_router.get("/")
async def root():
    return {"message": "Workspace Ruang Sanad API", "version": "1.2"}


@api_router.get("/tasks", response_model=List[Task])
async def list_tasks(
    kategori: Optional[str] = None,
    status: Optional[str] = None,
    list_id: Optional[str] = None,
    label_id: Optional[str] = None,
    divisi_id: Optional[str] = None,
    archived: bool = False,
    search: Optional[str] = None,
    tipe: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    scope = await user_scope(user)
    if archived:
        q: dict = {"archived": True}
    else:
        q = {"archived": {"$ne": True}}
    if kategori:
        q["kategori"] = kategori
    if status:
        q["status"] = status
    if list_id:
        q["list_id"] = list_id
    if label_id:
        q["label_ids"] = label_id
    if search:
        q["nama"] = {"$regex": search, "$options": "i"}
    if tipe == "PROJECT":
        q["kategori"] = "PROJECT"
    elif tipe == "RUTIN":
        q["kategori"] = {"$in": ["HARIAN", "MINGGUAN", "BULANAN"]}
    # Scope: anggota melihat tugas yang ditugaskan ke/dari dia (LINTAS divisi — delegasi),
    # plus tugas rutin di divisinya. JANGAN terapkan filter divisi_id top-level untuk anggota,
    # karena itu menyembunyikan tugas delegasi lintas divisi.
    if not scope["is_spv"]:
        my_ang = scope["anggota_id"]
        my_div = scope["divisi_id"]
        or_clauses = []
        if my_ang:
            or_clauses.append({"penerima_tugas_id": my_ang})
            or_clauses.append({"pemberi_id": my_ang})
        if my_div:
            or_clauses.append({
                "kategori": {"$in": ["HARIAN", "MINGGUAN", "BULANAN"]},
                "divisi_id": my_div,
            })
        if not or_clauses:
            return []
        q["$or"] = or_clauses
    elif divisi_id:
        q["divisi_id"] = divisi_id
    rows = await db.tasks.find(q, {"_id": 0}).sort([("urutan", 1), ("created_at", -1)]).to_list(3000)
    return rows


@api_router.post("/tasks", response_model=Task)
async def create_task(payload: TaskBase, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    data = payload.model_dump()
    # Anggota: auto-set pemberi_id = self anggota_id, and default penerima = self if empty
    if not scope["is_spv"]:
        if not scope["anggota_id"]:
            raise HTTPException(403, "Akun belum terhubung ke anggota. Hubungi SPV.")
        data["pemberi_id"] = scope["anggota_id"]
        # If penerima kosong, default = diri sendiri
        if not data.get("penerima_tugas_id"):
            data["penerima_tugas_id"] = scope["anggota_id"]
        # Tugas tetap "tinggal" di divisi & list pembuat (default Backlog) —
        # penerima lintas divisi tetap melihatnya via scope penerima_tugas_id.
        data["divisi_id"] = scope["divisi_id"]
    else:
        # SPV yang tertaut anggota tercatat sebagai pemberi agar tugas muncul sebagai delegasi darinya
        if not data.get("pemberi_id"):
            data["pemberi_id"] = scope["anggota_id"] or None
    task = Task(**data)
    await db.tasks.insert_one(task.model_dump())
    return task


@api_router.get("/tasks/{task_id}", response_model=Task)
async def get_task(task_id: str):
    doc = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Task not found")
    return doc


@api_router.put("/tasks/{task_id}", response_model=Task)
async def update_task(task_id: str, payload: TaskUpdate, user: dict = Depends(get_current_user)):
    task = await _assert_task_access(task_id, user)
    scope = await user_scope(user)
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    # Anggota: kalau hanya sebagai pemberi (bukan penerima) → read-only, tidak boleh update
    if not scope["is_spv"]:
        my_ang = scope["anggota_id"]
        is_penerima = my_ang and task.get("penerima_tugas_id") == my_ang
        is_pemberi = my_ang and task.get("pemberi_id") == my_ang
        if not is_penerima and is_pemberi:
            raise HTTPException(403, "Anda hanya pemberi tugas — read-only. Hanya penerima yang dapat mengubah tugas ini.")
        if "divisi_id" in update and update["divisi_id"] != scope["divisi_id"] and update["divisi_id"] != task.get("divisi_id"):
            raise HTTPException(403, "Anggota tidak dapat memindahkan tugas ke tim lain.")
    # Sinkron status otomatis saat list_id berubah (robust untuk semua client):
    # list is_done → SELESAI; list pertama (urutan<=1) → BELUM_MULAI; lainnya → DALAM_PROSES.
    if "list_id" in update and "status" not in update:
        if update["list_id"]:
            tl = await db.task_lists.find_one({"id": update["list_id"]}, {"_id": 0})
            if tl:
                update["status"] = "SELESAI" if tl.get("is_done") else ("BELUM_MULAI" if tl.get("urutan", 99) <= 1 else "DALAM_PROSES")
    update["updated_at"] = now_iso()
    result = await db.tasks.find_one_and_update(
        {"id": task_id},
        {"$set": update},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Task not found")
    return result


class RevisiPayload(BaseModel):
    catatan: str = Field(min_length=1)


@api_router.post("/tasks/{task_id}/revisi", response_model=Task)
async def revisi_task(task_id: str, payload: RevisiPayload, user: dict = Depends(get_current_user)):
    """Pemberi tugas (atau SPV) meminta revisi: set status=REVISI + catatan untuk penerima."""
    task = await _assert_task_access(task_id, user)
    scope = await user_scope(user)
    if not scope["is_spv"]:
        my_ang = scope["anggota_id"]
        if not my_ang or task.get("pemberi_id") != my_ang or task.get("penerima_tugas_id") == my_ang:
            raise HTTPException(403, "Hanya pemberi tugas yang dapat meminta revisi.")
    result = await db.tasks.find_one_and_update(
        {"id": task_id},
        {
            "$set": {"status": "REVISI", "revisi_catatan": payload.catatan, "revisi_at": now_iso(), "updated_at": now_iso()},
            "$inc": {"revisi_count": 1},
        },
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Task not found")
    return result


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, user: dict = Depends(get_current_user)):
    await _assert_task_access(task_id, user)
    r = await db.tasks.delete_one({"id": task_id})
    if r.deleted_count == 0:
        raise HTTPException(404, "Task not found")
    await db.todo_entries.delete_many({"task_id": task_id})
    return {"ok": True}


@api_router.post("/tasks/{task_id}/archive", response_model=Task)
async def archive_task(task_id: str, user: dict = Depends(get_current_user)):
    await _assert_task_access(task_id, user)
    result = await db.tasks.find_one_and_update(
        {"id": task_id},
        {"$set": {"archived": True, "archived_at": now_iso(), "updated_at": now_iso()}},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Task not found")
    return result


@api_router.post("/tasks/{task_id}/unarchive", response_model=Task)
async def unarchive_task(task_id: str, user: dict = Depends(get_current_user)):
    await _assert_task_access(task_id, user)
    result = await db.tasks.find_one_and_update(
        {"id": task_id},
        {"$set": {"archived": False, "archived_at": None, "updated_at": now_iso()}},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Task not found")
    return result


@api_router.post("/tasks/{task_id}/move", response_model=Task)
async def move_task(task_id: str, payload: MoveTaskPayload, user: dict = Depends(get_current_user)):
    task = await _assert_task_access(task_id, user)
    scope = await user_scope(user)
    if not scope["is_spv"] and payload.divisi_id and payload.divisi_id != scope["divisi_id"] and payload.divisi_id != task.get("divisi_id"):
        raise HTTPException(403, "Anggota tidak dapat memindahkan tugas ke tim lain.")
    # Pindah ke perorangan (reassign penerima) — boleh lintas divisi.
    # Hanya SPV, pemberi, atau penerima saat ini yang boleh.
    if payload.penerima_tugas_id is not None and payload.penerima_tugas_id != task.get("penerima_tugas_id"):
        my_ang = scope["anggota_id"]
        if not scope["is_spv"] and my_ang not in (task.get("pemberi_id"), task.get("penerima_tugas_id")):
            raise HTTPException(403, "Hanya pemberi/penerima tugas atau SPV yang bisa memindahkan ke orang lain.")
    update: dict = {"updated_at": now_iso()}
    if payload.penerima_tugas_id is not None and payload.penerima_tugas_id != task.get("penerima_tugas_id"):
        update["penerima_tugas_id"] = payload.penerima_tugas_id
        update["moved_at"] = now_iso()
    if payload.divisi_id is not None:
        update["divisi_id"] = payload.divisi_id
        if payload.divisi_id != task.get("divisi_id"):
            update["moved_at"] = now_iso()
    if payload.list_id is not None:
        update["list_id"] = payload.list_id
    if payload.urutan is not None:
        update["urutan"] = payload.urutan
        target_divisi = update.get("divisi_id", task.get("divisi_id"))
        target_list = update.get("list_id", task.get("list_id"))
        await db.tasks.update_many(
            {
                "id": {"$ne": task_id},
                "divisi_id": target_divisi,
                "list_id": target_list,
                "urutan": {"$gte": payload.urutan},
            },
            {"$inc": {"urutan": 1}},
        )
    result = await db.tasks.find_one_and_update(
        {"id": task_id}, {"$set": update},
        return_document=True, projection={"_id": 0},
    )
    return result


class BulkDeleteRequest(BaseModel):
    ids: List[str]


@api_router.post("/tasks/bulk_delete")
async def bulk_delete_tasks(payload: BulkDeleteRequest, user: dict = Depends(get_current_user)):
    if not payload.ids:
        return {"ok": True, "deleted": 0}
    await _assert_bulk_task_access(payload.ids, user)
    r = await db.tasks.delete_many({"id": {"$in": payload.ids}})
    await db.todo_entries.delete_many({"task_id": {"$in": payload.ids}})
    return {"ok": True, "deleted": r.deleted_count}


@api_router.post("/tasks/bulk_archive")
async def bulk_archive_tasks(payload: BulkIdsRequest, user: dict = Depends(get_current_user)):
    if not payload.ids:
        return {"ok": True, "archived": 0}
    await _assert_bulk_task_access(payload.ids, user)
    r = await db.tasks.update_many(
        {"id": {"$in": payload.ids}},
        {"$set": {"archived": True, "archived_at": now_iso(), "updated_at": now_iso()}},
    )
    return {"ok": True, "archived": r.modified_count}


@api_router.post("/tasks/bulk_unarchive")
async def bulk_unarchive_tasks(payload: BulkIdsRequest, user: dict = Depends(get_current_user)):
    if not payload.ids:
        return {"ok": True, "unarchived": 0}
    await _assert_bulk_task_access(payload.ids, user)
    r = await db.tasks.update_many(
        {"id": {"$in": payload.ids}},
        {"$set": {"archived": False, "archived_at": None, "updated_at": now_iso()}},
    )
    return {"ok": True, "unarchived": r.modified_count}


@api_router.post("/tasks/bulk_move")
async def bulk_move_tasks(payload: BulkMoveRequest, user: dict = Depends(get_current_user)):
    if not payload.ids:
        return {"ok": True, "moved": 0}
    scope = await user_scope(user)
    if not scope["is_spv"]:
        raise HTTPException(403, "Hanya SPV yang bisa memindahkan tugas antar tim.")
    update: dict = {"updated_at": now_iso()}
    if payload.divisi_id is not None:
        update["divisi_id"] = payload.divisi_id
        update["moved_at"] = now_iso()
    if payload.list_id is not None:
        update["list_id"] = payload.list_id
    r = await db.tasks.update_many({"id": {"$in": payload.ids}}, {"$set": update})
    return {"ok": True, "moved": r.modified_count}


@api_router.get("/task_unread")
async def unread_tasks(user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    q: dict = {"moved_at": {"$ne": None}, "archived": {"$ne": True}}
    if not scope["is_spv"]:
        if not scope["divisi_id"]:
            return {"total": 0, "by_divisi": {}}
        q["divisi_id"] = scope["divisi_id"]
    rows = await db.tasks.find(q, {"_id": 0, "id": 1, "divisi_id": 1, "moved_at": 1}).to_list(3000)
    by_divisi: dict = {}
    for r in rows:
        d = r.get("divisi_id") or "none"
        by_divisi[d] = by_divisi.get(d, 0) + 1
    return {"total": len(rows), "by_divisi": by_divisi}


# ============ NOTIFICATIONS (TUGAS MASUK DARI TIM LAIN) ============
@api_router.get("/notifications/incoming")
async def incoming_notifications(user: dict = Depends(get_current_user)):
    """Tugas yang didelegasikan KE saya oleh orang lain, sejak terakhir dilihat."""
    scope = await user_scope(user)
    if not scope["anggota_id"]:
        return {"count": 0, "items": []}
    seen = await db.user_task_seen.find_one({"user_id": scope["user_id"]}, {"_id": 0})
    q: dict = {
        "penerima_tugas_id": scope["anggota_id"],
        "pemberi_id": {"$nin": [None, scope["anggota_id"]]},
        "archived": {"$ne": True},
    }
    if seen and seen.get("last_seen_at"):
        q["created_at"] = {"$gt": seen["last_seen_at"]}
    rows = await db.tasks.find(
        q, {"_id": 0, "id": 1, "nama": 1, "pemberi_id": 1, "divisi_id": 1, "created_at": 1, "status": 1}
    ).sort("created_at", -1).to_list(50)
    pemberi_ids = [pid for pid in {r.get("pemberi_id") for r in rows} if pid]
    divisi_ids = [did for did in {r.get("divisi_id") for r in rows} if did]
    anggota_rows = await db.anggota.find({"id": {"$in": pemberi_ids}}, {"_id": 0, "id": 1, "nama": 1}).to_list(200) if pemberi_ids else []
    divisi_rows = await db.divisi.find({"id": {"$in": divisi_ids}}, {"_id": 0, "id": 1, "nama": 1}).to_list(200) if divisi_ids else []
    amap = {a["id"]: a["nama"] for a in anggota_rows}
    dmap = {d["id"]: d["nama"] for d in divisi_rows}
    items = [{**r, "pemberi_nama": amap.get(r.get("pemberi_id"), ""), "divisi_nama": dmap.get(r.get("divisi_id"), "")} for r in rows]
    return {"count": len(items), "items": items}


@api_router.post("/notifications/mark_seen")
async def mark_notifications_seen(user: dict = Depends(get_current_user)):
    await db.user_task_seen.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"last_seen_at": now_iso()}},
        upsert=True,
    )
    return {"ok": True}


@api_router.post("/task_mark_seen")
async def mark_seen_tasks(divisi_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    q: dict = {"moved_at": {"$ne": None}}
    if divisi_id:
        q["divisi_id"] = divisi_id
    if not scope["is_spv"]:
        if not scope["divisi_id"]:
            return {"ok": True, "cleared": 0}
        q["divisi_id"] = scope["divisi_id"]
    r = await db.tasks.update_many(q, {"$set": {"moved_at": None}})
    return {"ok": True, "cleared": r.modified_count}


@api_router.post("/tasks/reorder")
async def reorder_tasks(payload: ReorderRequest, user: dict = Depends(get_current_user)):
    await _assert_bulk_task_access(payload.task_ids, user)
    for i, tid in enumerate(payload.task_ids):
        await db.tasks.update_one({"id": tid}, {"$set": {"urutan": i}})
    return {"ok": True, "reordered": len(payload.task_ids)}


# ============ KATEGORI ROUTES ============
@api_router.get("/kategori", response_model=List[Kategori])
async def list_kategori():
    rows = await db.kategori.find({}, {"_id": 0}).sort("urutan", 1).to_list(500)
    return rows


@api_router.post("/kategori", response_model=Kategori)
async def create_kategori(payload: KategoriCreate):
    count = await db.kategori.count_documents({})
    item = Kategori(**payload.model_dump(), urutan=count + 1)
    await db.kategori.insert_one(item.model_dump())
    return item


@api_router.put("/kategori/{kategori_id}", response_model=Kategori)
async def update_kategori(kategori_id: str, payload: KategoriCreate):
    result = await db.kategori.find_one_and_update(
        {"id": kategori_id}, {"$set": payload.model_dump()},
        return_document=True, projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Kategori not found")
    return result


@api_router.delete("/kategori/{kategori_id}")
async def delete_kategori(kategori_id: str):
    r = await db.kategori.delete_one({"id": kategori_id})
    await db.tasks.update_many({"kategori_id": kategori_id}, {"$set": {"kategori_id": None}})
    return {"ok": True, "deleted": r.deleted_count}


# ============ ANGGOTA ROUTES ============
@api_router.get("/anggota", response_model=List[Anggota])
async def list_anggota(divisi_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    q: dict = {}
    if divisi_id:
        q["divisi_id"] = divisi_id
    # Semua user (termasuk anggota) melihat daftar anggota SEMUA divisi —
    # dibutuhkan untuk delegasi lintas divisi (picker penerima tugas). Read-only nama.
    rows = await db.anggota.find(q, {"_id": 0}).sort("urutan", 1).to_list(1000)
    return rows


@api_router.post("/anggota", response_model=Anggota)
async def create_anggota(payload: AnggotaCreate, _: dict = Depends(require_spv)):
    count = await db.anggota.count_documents({"divisi_id": payload.divisi_id})
    item = Anggota(**payload.model_dump(), urutan=count + 1)
    await db.anggota.insert_one(item.model_dump())
    return item


@api_router.get("/anggota/analytics")
async def anggota_analytics(divisi_id: str, month: Optional[str] = None, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    if not scope["is_spv"]:
        if not scope["divisi_id"] or scope["divisi_id"] != divisi_id:
            raise HTTPException(403, "Analytics tim lain tidak dapat diakses.")
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    year, mm = map(int, month.split("-"))
    nxt = f"{year+1}-01" if mm == 12 else f"{year}-{str(mm+1).zfill(2)}"
    start_iso = f"{month}-01T00:00:00"
    end_iso = f"{nxt}-01T00:00:00"

    anggota = await db.anggota.find({"divisi_id": divisi_id}, {"_id": 0}).sort("urutan", 1).to_list(500)
    result = []
    for a in anggota:
        base = {"divisi_id": divisi_id, "penerima_tugas_id": a["id"], "created_at": {"$gte": start_iso, "$lt": end_iso}}
        total = await db.tasks.count_documents(base)
        selesai = await db.tasks.count_documents({**base, "status": "SELESAI"})
        terkendala = await db.tasks.count_documents({**base, "status": "TERKENDALA"})
        result.append({**a, "total": total, "selesai": selesai, "terkendala": terkendala})
    result.sort(key=lambda x: x["selesai"], reverse=True)
    return {"month": month, "anggota": result}


@api_router.get("/anggota/{anggota_id}")
async def get_anggota(anggota_id: str):
    doc = await db.anggota.find_one({"id": anggota_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Anggota not found")
    return doc


@api_router.put("/anggota/{anggota_id}", response_model=Anggota)
async def update_anggota(anggota_id: str, payload: AnggotaCreate, _: dict = Depends(require_spv)):
    result = await db.anggota.find_one_and_update(
        {"id": anggota_id}, {"$set": payload.model_dump()},
        return_document=True, projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Anggota not found")
    return result


@api_router.delete("/anggota/{anggota_id}")
async def delete_anggota(anggota_id: str, _: dict = Depends(require_spv)):
    r = await db.anggota.delete_one({"id": anggota_id})
    await db.tasks.update_many({"penerima_tugas_id": anggota_id}, {"$set": {"penerima_tugas_id": None}})
    return {"ok": True, "deleted": r.deleted_count}


# ============ DIVISI (TIM/WORKSPACE) ROUTES ============
@api_router.get("/divisi", response_model=List[Divisi])
async def list_divisi(user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    rows = await db.divisi.find({}, {"_id": 0}).sort("urutan", 1).to_list(200)
    if not rows:
        # Auto-migrate from existing tasks.divisi string values
        unique = await db.tasks.distinct("divisi")
        names = [n for n in unique if n and isinstance(n, str)]
        if not names:
            names = ["Umum"]
        created: list = []
        for i, name in enumerate(names):
            d = Divisi(nama=name, urutan=i + 1)
            await db.divisi.insert_one(d.model_dump())
            created.append(d)
            await db.tasks.update_many(
                {"divisi": name, "divisi_id": {"$exists": False}},
                {"$set": {"divisi_id": d.id}},
            )
        first_id = created[0].id
        # Existing lists without divisi_id → assign to first divisi
        await db.task_lists.update_many(
            {"divisi_id": {"$exists": False}}, {"$set": {"divisi_id": first_id}}
        )
        await db.task_lists.update_many(
            {"divisi_id": None}, {"$set": {"divisi_id": first_id}}
        )
        # Seed default lists for other divisi that have none
        for d in created[1:]:
            count = await db.task_lists.count_documents({"divisi_id": d.id})
            if count == 0:
                defaults = [
                    {"nama": "Backlog", "warna": "#94a3b8", "urutan": 1, "is_done": False},
                    {"nama": "Dikerjakan", "warna": "#f59e0b", "urutan": 2, "is_done": False},
                    {"nama": "Selesai", "warna": "#10b981", "urutan": 3, "is_done": True},
                ]
                for x in defaults:
                    tl = TaskList(**x, divisi_id=d.id)
                    await db.task_lists.insert_one(tl.model_dump())
        rows = await db.divisi.find({}, {"_id": 0}).sort("urutan", 1).to_list(200)
    # Semua user melihat semua divisi (read-only nama) — dibutuhkan untuk picker delegasi.
    # Frontend membatasi TAB workspace anggota ke divisinya sendiri.
    return rows


@api_router.post("/divisi", response_model=Divisi)
async def create_divisi(payload: DivisiCreate, _: dict = Depends(require_spv)):
    count = await db.divisi.count_documents({})
    item = Divisi(**payload.model_dump(), urutan=count + 1)
    await db.divisi.insert_one(item.model_dump())
    defaults = [
        {"nama": "Backlog", "warna": "#94a3b8", "urutan": 1, "is_done": False},
        {"nama": "Dikerjakan", "warna": "#f59e0b", "urutan": 2, "is_done": False},
        {"nama": "Selesai", "warna": "#10b981", "urutan": 3, "is_done": True},
    ]
    for d in defaults:
        tl = TaskList(**d, divisi_id=item.id)
        await db.task_lists.insert_one(tl.model_dump())
    return item


@api_router.put("/divisi/{divisi_id}", response_model=Divisi)
async def update_divisi(divisi_id: str, payload: DivisiCreate, _: dict = Depends(require_spv)):
    result = await db.divisi.find_one_and_update(
        {"id": divisi_id}, {"$set": payload.model_dump()},
        return_document=True, projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Divisi not found")
    return result


@api_router.delete("/divisi/{divisi_id}")
async def delete_divisi(divisi_id: str, _: dict = Depends(require_spv)):
    remaining = await db.divisi.find(
        {"id": {"$ne": divisi_id}}, {"_id": 0}
    ).sort("urutan", 1).to_list(1)
    if not remaining:
        raise HTTPException(400, "Tidak bisa hapus satu-satunya divisi")
    target = remaining[0]["id"]
    await db.tasks.update_many({"divisi_id": divisi_id}, {"$set": {"divisi_id": target}})
    await db.task_lists.delete_many({"divisi_id": divisi_id})
    r = await db.divisi.delete_one({"id": divisi_id})
    return {"ok": True, "deleted": r.deleted_count}


# ============ TASK LISTS ROUTES ============
@api_router.get("/task_lists", response_model=List[TaskList])
async def list_task_lists(divisi_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    if not scope["is_spv"]:
        if scope["divisi_id"]:
            divisi_id = scope["divisi_id"]
        else:
            return []
    q: dict = {}
    if divisi_id:
        q["divisi_id"] = divisi_id
    rows = await db.task_lists.find(q, {"_id": 0}).sort("urutan", 1).to_list(500)
    if not rows and divisi_id:
        defaults = [
            {"nama": "Backlog", "warna": "#94a3b8", "urutan": 1, "is_done": False},
            {"nama": "Dikerjakan", "warna": "#f59e0b", "urutan": 2, "is_done": False},
            {"nama": "Selesai", "warna": "#10b981", "urutan": 3, "is_done": True},
        ]
        for d in defaults:
            item = TaskList(**d, divisi_id=divisi_id)
            await db.task_lists.insert_one(item.model_dump())
        rows = await db.task_lists.find(q, {"_id": 0}).sort("urutan", 1).to_list(500)
    return rows


@api_router.post("/task_lists", response_model=TaskList)
async def create_task_list(payload: TaskListCreate):
    data = payload.model_dump()
    if not data.get("divisi_id"):
        first = await db.divisi.find_one({}, {"_id": 0}, sort=[("urutan", 1)])
        if first:
            data["divisi_id"] = first["id"]
    count = await db.task_lists.count_documents({"divisi_id": data.get("divisi_id")})
    if data.get("urutan") is None:
        data["urutan"] = count + 1
    item = TaskList(**data)
    await db.task_lists.insert_one(item.model_dump())
    return item


@api_router.put("/task_lists/{list_id}", response_model=TaskList)
async def update_task_list(list_id: str, payload: TaskListCreate):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    result = await db.task_lists.find_one_and_update(
        {"id": list_id},
        {"$set": update},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "List not found")
    return result


@api_router.delete("/task_lists/{list_id}")
async def delete_task_list(list_id: str):
    r = await db.task_lists.delete_one({"id": list_id})
    await db.tasks.update_many({"list_id": list_id}, {"$set": {"list_id": None}})
    return {"ok": True, "deleted": r.deleted_count}


# ============ TASK LABELS ROUTES ============
@api_router.get("/task_labels", response_model=List[TaskLabel])
async def list_task_labels():
    rows = await db.task_labels.find({}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return rows


@api_router.post("/task_labels", response_model=TaskLabel)
async def create_task_label(payload: TaskLabelCreate):
    item = TaskLabel(**payload.model_dump())
    await db.task_labels.insert_one(item.model_dump())
    return item


@api_router.put("/task_labels/{label_id}", response_model=TaskLabel)
async def update_task_label(label_id: str, payload: TaskLabelCreate):
    result = await db.task_labels.find_one_and_update(
        {"id": label_id},
        {"$set": payload.model_dump()},
        return_document=True,
        projection={"_id": 0},
    )
    if not result:
        raise HTTPException(404, "Label not found")
    return result


@api_router.delete("/task_labels/{label_id}")
async def delete_task_label(label_id: str):
    r = await db.task_labels.delete_one({"id": label_id})
    await db.tasks.update_many({}, {"$pull": {"label_ids": label_id}})
    return {"ok": True, "deleted": r.deleted_count}


# ============ TODO TRACKER (RUTIN) ROUTES ============
@api_router.get("/todo/entries", response_model=List[TodoEntry])
async def list_todo_entries(
    start: Optional[str] = None,
    end: Optional[str] = None,
    task_ids: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    scope = await user_scope(user)
    q: dict = {}
    if start or end:
        q["period"] = {}
        if start:
            q["period"]["$gte"] = start
        if end:
            q["period"]["$lte"] = end
    if task_ids:
        q["task_id"] = {"$in": task_ids.split(",")}
    if not scope["is_spv"]:
        q["user_id"] = scope["user_id"]
    rows = await db.todo_entries.find(q, {"_id": 0}).to_list(20000)
    return rows


@api_router.post("/todo/entries", response_model=TodoEntry)
async def upsert_todo_entry(payload: TodoEntryCreate, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    filter_q = {"task_id": payload.task_id, "period": payload.period}
    if not scope["is_spv"]:
        filter_q["user_id"] = scope["user_id"]
    existing = await db.todo_entries.find_one(filter_q, {"_id": 0})
    if existing:
        await db.todo_entries.update_one(
            {"id": existing["id"]}, {"$set": {"checked": payload.checked}}
        )
        existing["checked"] = payload.checked
        return existing
    entry = TodoEntry(**payload.model_dump())
    doc = entry.model_dump()
    doc["user_id"] = scope["user_id"]
    await db.todo_entries.insert_one(doc)
    return doc


# ============ AMALIYAH ITEMS ROUTES ============
@api_router.get("/amaliyah/items", response_model=List[AmaliyahItem])
async def list_amaliyah_items():
    rows = await db.amaliyah_items.find({}, {"_id": 0}).sort("urutan", 1).to_list(500)
    if not rows:
        # seed default
        defaults = [
            {"nama": "Sedekah Subuh", "target_metrik": "1x/hari", "urutan": 1},
            {"nama": "Tilawah 1 Lembar", "target_metrik": "1 lembar/hari", "urutan": 2},
            {"nama": "Dzikir Pagi", "target_metrik": "1x/hari", "urutan": 3},
            {"nama": "Dzikir Petang", "target_metrik": "1x/hari", "urutan": 4},
            {"nama": "Tahajud", "target_metrik": "1x/hari", "urutan": 5},
        ]
        for d in defaults:
            item = AmaliyahItem(**d)
            await db.amaliyah_items.insert_one(item.model_dump())
        rows = await db.amaliyah_items.find({}, {"_id": 0}).sort("urutan", 1).to_list(500)
    return rows


@api_router.post("/amaliyah/items", response_model=AmaliyahItem)
async def create_amaliyah_item(payload: AmaliyahItemCreate, _: dict = Depends(require_spv)):
    count = await db.amaliyah_items.count_documents({})
    item = AmaliyahItem(**payload.model_dump(), urutan=count + 1)
    await db.amaliyah_items.insert_one(item.model_dump())
    return item


@api_router.put("/amaliyah/items/{item_id}", response_model=AmaliyahItem)
async def update_amaliyah_item(item_id: str, payload: AmaliyahItemCreate, _: dict = Depends(require_spv)):
    update = payload.model_dump()
    result = await db.amaliyah_items.find_one_and_update(
        {"id": item_id}, {"$set": update}, return_document=True, projection={"_id": 0}
    )
    if not result:
        raise HTTPException(404, "Item not found")
    return result


@api_router.delete("/amaliyah/items/{item_id}")
async def delete_amaliyah_item(item_id: str, _: dict = Depends(require_spv)):
    r = await db.amaliyah_items.delete_one({"id": item_id})
    await db.amaliyah_entries.delete_many({"item_id": item_id})
    return {"ok": True, "deleted": r.deleted_count}


class AmaliyahReorder(BaseModel):
    item_ids: List[str]


@api_router.post("/amaliyah/items/reorder")
async def reorder_amaliyah_items(payload: AmaliyahReorder, _: dict = Depends(require_spv)):
    for i, iid in enumerate(payload.item_ids):
        await db.amaliyah_items.update_one({"id": iid}, {"$set": {"urutan": i + 1}})
    return {"ok": True, "reordered": len(payload.item_ids)}


@api_router.post("/amaliyah/items/bulk_delete")
async def bulk_delete_amaliyah_items(payload: BulkIdsRequest, _: dict = Depends(require_spv)):
    if not payload.ids:
        return {"ok": True, "deleted": 0}
    r = await db.amaliyah_items.delete_many({"id": {"$in": payload.ids}})
    await db.amaliyah_entries.delete_many({"item_id": {"$in": payload.ids}})
    return {"ok": True, "deleted": r.deleted_count}


@api_router.get("/amaliyah/streak")
async def amaliyah_streak(user: dict = Depends(get_current_user)):
    """Streak = consecutive days ending today where user checked >=1 amaliyah.
    Returns streak count + longest + earned badges (7/14/30/60/100/180/365)."""
    scope = await user_scope(user)
    from datetime import timedelta as _td
    today = date.today()
    # Fetch last 400 days of checked entries
    start = (today - _td(days=400)).isoformat()
    rows = await db.amaliyah_entries.find(
        {"user_id": scope["user_id"], "checked": True, "tanggal": {"$gte": start}},
        {"_id": 0, "tanggal": 1},
    ).to_list(20000)
    checked_days = set(r["tanggal"] for r in rows)
    # current streak
    current = 0
    d = today
    while d.isoformat() in checked_days:
        current += 1
        d = d - _td(days=1)
    # longest streak (walk backwards)
    longest = 0
    run = 0
    d = today
    for _ in range(400):
        if d.isoformat() in checked_days:
            run += 1
            longest = max(longest, run)
        else:
            run = 0
        d = d - _td(days=1)
    # Total checked distinct days
    total_days = len(checked_days)
    tiers = [7, 14, 30, 60, 100, 180, 365]
    badges = [t for t in tiers if longest >= t]
    next_target = next((t for t in tiers if t > longest), None)
    return {
        "current_streak": current,
        "longest_streak": longest,
        "total_days": total_days,
        "badges": badges,
        "next_target": next_target,
    }


# ============ AMALIYAH ENTRIES ROUTES ============
@api_router.get("/amaliyah/entries", response_model=List[AmaliyahEntry])
async def list_entries(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    q: dict = {}
    if start or end:
        q["tanggal"] = {}
        if start:
            q["tanggal"]["$gte"] = start
        if end:
            q["tanggal"]["$lte"] = end
    # Amaliyah is strictly personal — each user sees only their own entries.
    q["user_id"] = scope["user_id"]
    rows = await db.amaliyah_entries.find(q, {"_id": 0}).to_list(5000)
    return rows


@api_router.post("/amaliyah/entries", response_model=AmaliyahEntry)
async def upsert_entry(payload: AmaliyahEntryCreate, user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    existing = await db.amaliyah_entries.find_one(
        {"item_id": payload.item_id, "tanggal": payload.tanggal, "user_id": scope["user_id"]},
        {"_id": 0},
    )
    if existing:
        await db.amaliyah_entries.update_one(
            {"id": existing["id"]},
            {"$set": {"checked": payload.checked, "catatan": payload.catatan or ""}},
        )
        existing["checked"] = payload.checked
        existing["catatan"] = payload.catatan or ""
        return existing
    entry = AmaliyahEntry(**payload.model_dump(), user_id=scope["user_id"])
    await db.amaliyah_entries.insert_one(entry.model_dump())
    return entry


# ============ RAPORT ROUTES ============
async def _raport_summary_impl(start: Optional[str], end: Optional[str], user: dict, anggota_id: Optional[str] = None) -> dict:
    scope = await user_scope(user)
    # Task query
    task_query: dict = {}
    if start or end:
        task_query["created_at"] = {}
        if start:
            task_query["created_at"]["$gte"] = start
        if end:
            task_query["created_at"]["$lte"] = end + "T23:59:59"

    all_tasks = await db.tasks.find(task_query, {"_id": 0}).to_list(5000)
    if not scope["is_spv"]:
        all_tasks = [t for t in all_tasks if t.get("divisi_id") == scope["divisi_id"]] if scope["divisi_id"] else []

    ang = None
    div_nama = None
    if anggota_id:
        ang = await db.anggota.find_one({"id": anggota_id}, {"_id": 0})
        if ang:
            all_tasks = [t for t in all_tasks if t.get("penerima_tugas_id") == anggota_id]
            div = await db.divisi.find_one({"id": ang.get("divisi_id")}, {"_id": 0}) if ang.get("divisi_id") else None
            div_nama = div.get("nama") if div else None

    total = len(all_tasks)
    selesai = sum(1 for t in all_tasks if t.get("status") == "SELESAI")
    dalam_proses = sum(1 for t in all_tasks if t.get("status") == "DALAM_PROSES")
    terkendala = sum(1 for t in all_tasks if t.get("status") == "TERKENDALA")
    belum_mulai = sum(1 for t in all_tasks if t.get("status") == "BELUM_MULAI")
    today = date.today().isoformat()
    overdue = sum(1 for t in all_tasks if t.get("deadline") and t.get("deadline") < today and t.get("status") != "SELESAI")
    task_score = round((selesai / total * 100) if total else 0, 1)

    # Amaliyah
    entry_q: dict = {"checked": True}
    if start or end:
        entry_q["tanggal"] = {}
        if start:
            entry_q["tanggal"]["$gte"] = start
        if end:
            entry_q["tanggal"]["$lte"] = end
    if anggota_id and ang and ang.get("user_id"):
        entry_q["user_id"] = ang["user_id"]
    elif anggota_id:
        # anggota belum di-link → no entries
        entry_q["user_id"] = "__none__"
    elif not scope["is_spv"]:
        entry_q["user_id"] = scope["user_id"]
    entries = await db.amaliyah_entries.count_documents(entry_q)
    items_count = await db.amaliyah_items.count_documents({})

    days = 30
    if start and end:
        try:
            d1 = datetime.fromisoformat(start).date()
            d2 = datetime.fromisoformat(end).date()
            days = max((d2 - d1).days + 1, 1)
        except Exception:
            days = 30
    target = items_count * days if items_count else 1
    amaliyah_score = round((entries / target * 100) if target else 0, 1)
    combined = round((task_score * 0.6 + amaliyah_score * 0.4), 1)
    if combined >= 80:
        auto_rec = "REWARD"
    elif combined < 50:
        auto_rec = "EVALUASI"
    else:
        auto_rec = "NETRAL"

    note_key = {"id": f"anggota:{anggota_id}"} if anggota_id else {"id": "singleton"}
    note = await db.raport_notes.find_one(note_key, {"_id": 0}) or {
        **note_key,
        "catatan_spv": "",
        "rekomendasi": "NETRAL",
        "updated_at": now_iso(),
    }

    result = {
        "task": {"total": total, "selesai": selesai, "dalam_proses": dalam_proses, "terkendala": terkendala,
                 "belum_mulai": belum_mulai, "overdue": overdue, "score": task_score},
        "amaliyah": {"total_entries": entries, "target": target, "items_count": items_count, "days": days, "score": amaliyah_score},
        "combined_score": combined,
        "auto_rekomendasi": auto_rec,
        "spv_note": note,
    }
    if anggota_id and ang:
        result["anggota"] = {"id": ang["id"], "nama": ang.get("nama"), "divisi_nama": div_nama}
        # Include task list for PDF
        result["tasks_list"] = sorted(all_tasks, key=lambda x: (x.get("deadline") or "9999", x.get("nama") or ""))
    return result


@api_router.get("/raport/summary")
async def raport_summary(start: Optional[str] = None, end: Optional[str] = None, anggota_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    # RBAC: anggota_id filter is SPV-only (or self, if anggota is linked to current user)
    if anggota_id:
        scope = await user_scope(user)
        if not scope["is_spv"]:
            own = await db.anggota.find_one({"id": anggota_id, "user_id": user.get("user_id")}, {"_id": 0})
            if not own:
                raise HTTPException(403, "Hanya SPV yang bisa melihat raport anggota lain.")
    return await _raport_summary_impl(start, end, user, anggota_id)


@api_router.put("/raport/note")
async def update_raport_note(payload: RaportNoteUpdate, anggota_id: Optional[str] = None, _: dict = Depends(require_spv)):
    key_id = f"anggota:{anggota_id}" if anggota_id else "singleton"
    doc = {
        "id": key_id,
        "catatan_spv": payload.catatan_spv,
        "rekomendasi": payload.rekomendasi,
        "updated_at": now_iso(),
    }
    await db.raport_notes.update_one({"id": key_id}, {"$set": doc}, upsert=True)
    return doc


# ============ DASHBOARD DIGEST (SPV) ============
@api_router.get("/dashboard/digest")
async def dashboard_digest(user: dict = Depends(get_current_user)):
    scope = await user_scope(user)
    today = date.today().isoformat()
    week_ago = (date.today() - timedelta(days=7)).isoformat()
    week_ahead = (date.today() + timedelta(days=3)).isoformat()

    base = {"archived": {"$ne": True}, "status": {"$ne": "SELESAI"}}
    if not scope["is_spv"] and scope["divisi_id"]:
        base["divisi_id"] = scope["divisi_id"]
    elif not scope["is_spv"]:
        return {"overdue": [], "today": [], "upcoming": [], "stagnant": [], "counts": {"overdue": 0, "today": 0, "upcoming": 0, "stagnant": 0}}

    overdue = await db.tasks.find({**base, "deadline": {"$lt": today, "$ne": None}}, {"_id": 0}).sort("deadline", 1).limit(50).to_list(50)
    today_t = await db.tasks.find({**base, "deadline": today}, {"_id": 0}).limit(50).to_list(50)
    upcoming = await db.tasks.find({**base, "deadline": {"$gt": today, "$lte": week_ahead}}, {"_id": 0}).sort("deadline", 1).limit(50).to_list(50)

    cutoff = (datetime.now(timezone.utc) - timedelta(days=3)).isoformat()
    stg_q = {**base, "status": {"$in": ["BELUM_MULAI", "DALAM_PROSES", "TERKENDALA"]}, "updated_at": {"$lt": cutoff}}
    stagnant = await db.tasks.find(stg_q, {"_id": 0}).sort("updated_at", 1).limit(30).to_list(30)

    ang_list = await db.anggota.find({}, {"_id": 0}).to_list(500)
    ang_map = {a["id"]: a["nama"] for a in ang_list}
    div_list = await db.divisi.find({}, {"_id": 0}).to_list(200)
    div_map = {d["id"]: d["nama"] for d in div_list}

    def _dec(rows):
        return [{**r, "penerima_nama": ang_map.get(r.get("penerima_tugas_id"), r.get("penerima_tugas") or "-"),
                 "divisi_nama": div_map.get(r.get("divisi_id"), "-")} for r in rows]

    return {
        "overdue": _dec(overdue),
        "today": _dec(today_t),
        "upcoming": _dec(upcoming),
        "stagnant": _dec(stagnant),
        "counts": {"overdue": len(overdue), "today": len(today_t), "upcoming": len(upcoming), "stagnant": len(stagnant)},
    }


# ============ EXCEL IMPORT ============
def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _map_kategori(val: str) -> str:
    v = _norm(val)
    if "harian" in v:
        return "HARIAN"
    if "mingguan" in v:
        return "MINGGUAN"
    if "bulanan" in v:
        return "BULANAN"
    if "project" in v or "proyek" in v:
        return "PROJECT"
    return "HARIAN"


def _map_status(val: str) -> str:
    v = _norm(val)
    if "selesai" in v or "done" in v:
        return "SELESAI"
    if "proses" in v or "progress" in v:
        return "DALAM_PROSES"
    if "kendala" in v or "block" in v:
        return "TERKENDALA"
    return "BELUM_MULAI"


def _cell_to_iso(val: Any) -> Optional[str]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    try:
        # excel serial number
        n = float(val)
        base = datetime(1899, 12, 30)
        from datetime import timedelta

        return (base + timedelta(days=n)).date().isoformat()
    except Exception:
        return str(val)


@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), _: dict = Depends(require_spv)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Gagal baca Excel: {e}")

    imported_tasks = 0
    imported_amaliyah = 0
    log: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find header row (first row with recognizable header)
        header_row_idx = None
        headers: List[str] = []
        for r_idx in range(1, min(ws.max_row + 1, 20)):
            row_vals = [str(c.value or "").strip().lower() for c in ws[r_idx]]
            joined = " ".join(row_vals)
            if any(k in joined for k in ["nama tugas", "initiative", "to do", "todo"]):
                header_row_idx = r_idx
                headers = row_vals
                break
            if "nama amalan" in joined or "amaliyah" in joined or "habit" in joined:
                header_row_idx = r_idx
                headers = row_vals
                break

        if header_row_idx is None:
            continue

        # detect type
        is_amaliyah = any(
            "amalan" in h or "habit" in h or "amaliyah" in h for h in headers
        )

        # Build header index map
        def find_col(*keys: str) -> Optional[int]:
            for i, h in enumerate(headers):
                for k in keys:
                    if k in h:
                        return i
            return None

        if is_amaliyah:
            col_nama = find_col("nama amalan", "habit", "amaliyah", "amalan")
            col_target = find_col("target")
            col_ket = find_col("keterangan", "catatan")
            for r_idx in range(header_row_idx + 1, ws.max_row + 1):
                row = [c.value for c in ws[r_idx]]
                if col_nama is None or col_nama >= len(row):
                    continue
                nama = row[col_nama]
                if not nama or not str(nama).strip():
                    continue
                target = (
                    str(row[col_target])
                    if col_target is not None
                    and col_target < len(row)
                    and row[col_target]
                    else "1x/hari"
                )
                ket = (
                    str(row[col_ket])
                    if col_ket is not None and col_ket < len(row) and row[col_ket]
                    else ""
                )
                # Deduplicate by name
                exists = await db.amaliyah_items.find_one({"nama": str(nama).strip()})
                if exists:
                    continue
                count = await db.amaliyah_items.count_documents({})
                item = AmaliyahItem(
                    nama=str(nama).strip(),
                    target_metrik=target,
                    keterangan=ket,
                    urutan=count + 1,
                )
                await db.amaliyah_items.insert_one(item.model_dump())
                imported_amaliyah += 1
        else:
            col_nama = find_col("nama tugas", "initiative", "to do", "todo")
            col_kat = find_col("kategori", "frekuensi")
            col_freq = find_col("frekuensi")
            col_status = find_col("status")
            col_pemberi = find_col("pemberi")
            col_mulai = find_col("waktu mulai", "start", "mulai")
            col_deadline = find_col("batas waktu", "deadline")
            col_catatan_tim = find_col("catatan tim", "hambatan")
            col_catatan_spv = find_col("catatan spv", "arahan")
            col_link = find_col("link", "output")

            for r_idx in range(header_row_idx + 1, ws.max_row + 1):
                row = [c.value for c in ws[r_idx]]
                if col_nama is None or col_nama >= len(row):
                    continue
                nama = row[col_nama]
                if not nama or not str(nama).strip():
                    continue
                nama_str = str(nama).strip()
                if len(nama_str) < 2:
                    continue

                def rv(i):
                    return row[i] if i is not None and i < len(row) else None

                task = TaskBase(
                    nama=nama_str,
                    kategori=_map_kategori(rv(col_kat) or rv(col_freq) or "HARIAN"),
                    frekuensi=str(rv(col_freq) or "RUTIN").upper()[:30],
                    status=_map_status(rv(col_status) or ""),
                    pemberi_tugas=str(rv(col_pemberi) or ""),
                    tanggal_mulai=_cell_to_iso(rv(col_mulai)),
                    deadline=_cell_to_iso(rv(col_deadline)),
                    catatan_tim=str(rv(col_catatan_tim) or ""),
                    catatan_spv=str(rv(col_catatan_spv) or ""),
                    link_output=str(rv(col_link) or ""),
                    divisi=sheet_name,
                )
                t = Task(**task.model_dump())
                await db.tasks.insert_one(t.model_dump())
                imported_tasks += 1

        log.append(
            f"Sheet '{sheet_name}': {'amaliyah' if is_amaliyah else 'tasks'} imported"
        )

    return {
        "ok": True,
        "imported_tasks": imported_tasks,
        "imported_amaliyah": imported_amaliyah,
        "log": log,
    }


# ============ RAPORT PDF EXPORT ============
@api_router.get("/raport/export.pdf")
async def export_raport_pdf(
    start: Optional[str] = None,
    end: Optional[str] = None,
    anggota_id: Optional[str] = None,
    user: dict = Depends(require_spv),
):
    summary = await _raport_summary_impl(start, end, user, anggota_id)
    ang = summary.get("anggota") or {}
    if anggota_id:
        subject = f"Raport Individu — {ang.get('nama', '')}"
        anggota_nama = ang.get("nama")
        divisi_nama = ang.get("divisi_nama")
        slug = (ang.get("nama") or "anggota").lower().replace(" ", "-")
    else:
        subject = "Raport Kinerja Tim"
        anggota_nama = None
        divisi_nama = None
        slug = "tim"
    pdf_bytes = build_raport_pdf(
        summary, start or "-", end or "-",
        subject=subject, anggota_nama=anggota_nama, divisi_nama=divisi_nama,
    )
    fname = f"raport-ruang-sanad-{slug}-{(start or 'all')}_{(end or 'all')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ============ GLOBAL SEARCH ============
@api_router.get("/search")
async def global_search(q: str = "", user: dict = Depends(get_current_user)):
    q = (q or "").strip()
    if len(q) < 2:
        return {"tasks": [], "amaliyah": [], "anggota": [], "divisi": []}
    scope = await user_scope(user)
    regex = {"$regex": q, "$options": "i"}

    # Tasks
    task_q = {"nama": regex, "archived": {"$ne": True}}
    if not scope["is_spv"]:
        if scope["divisi_id"]:
            task_q["divisi_id"] = scope["divisi_id"]
        else:
            task_q["_never"] = True
    tasks = await db.tasks.find(task_q, {"_id": 0, "id": 1, "nama": 1, "status": 1, "divisi_id": 1, "deadline": 1}).limit(20).to_list(20)

    # Amaliyah items (shared list)
    amal = await db.amaliyah_items.find({"nama": regex}, {"_id": 0, "id": 1, "nama": 1, "target_metrik": 1}).limit(10).to_list(10)

    # Anggota (peers in own divisi, or all for SPV)
    ang_q = {"nama": regex}
    if not scope["is_spv"]:
        if scope["divisi_id"]:
            ang_q["divisi_id"] = scope["divisi_id"]
        else:
            ang_q["_never"] = True
    anggota = await db.anggota.find(ang_q, {"_id": 0, "id": 1, "nama": 1, "divisi_id": 1, "warna": 1}).limit(10).to_list(10)

    # Divisi (SPV only sees all; anggota only theirs)
    div_q = {"nama": regex}
    divisi = await db.divisi.find(div_q, {"_id": 0, "id": 1, "nama": 1, "warna": 1}).limit(10).to_list(10)
    if not scope["is_spv"]:
        divisi = [d for d in divisi if d["id"] == scope["divisi_id"]] if scope["divisi_id"] else []

    # Decorate tasks with divisi_nama
    d_map = {d["id"]: d for d in await db.divisi.find({}, {"_id": 0, "id": 1, "nama": 1}).to_list(200)}
    tasks = [{**t, "divisi_nama": d_map.get(t.get("divisi_id"), {}).get("nama", "-")} for t in tasks]

    return {"tasks": tasks, "amaliyah": amal, "anggota": anggota, "divisi": divisi}


# Attach auth & monitoring routers to /api
api_router.include_router(build_auth_router(db))
api_router.include_router(build_monitoring_router(db, require_spv))
api_router.include_router(build_strategy_router(db, get_current_user, require_spv, user_scope))

# Include router
app.include_router(api_router)

_cors_env = os.environ.get('CORS_ORIGINS', '*').strip()
if _cors_env in ("", "*"):
    _cors_origins = ["*"]
    _allow_credentials = False  # credentials + wildcard is invalid
else:
    _cors_origins = [o.strip() for o in _cors_env.split(',') if o.strip()]
    _allow_credentials = True

# Always allow the configured frontend host + credentials when available
_frontend_url = os.environ.get('FRONTEND_URL', '').strip()
if _frontend_url and _frontend_url not in _cors_origins:
    _cors_origins.append(_frontend_url)
    _allow_credentials = True

app.add_middleware(
    CORSMiddleware,
    allow_credentials=_allow_credentials,
    allow_origins=_cors_origins,
    allow_origin_regex=r"https://.*\.preview\.emergentagent\.com",
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup():
    await seed_admin(db)
    # Migrate existing amaliyah_entries without user_id → assign to first SPV
    spv = await db.users.find_one({"role": "spv", "status": "approved"}, {"_id": 0})
    if spv:
        await db.amaliyah_entries.update_many(
            {"user_id": {"$in": [None, "", False]}}, {"$set": {"user_id": spv["user_id"]}}
        )
        await db.amaliyah_entries.update_many(
            {"user_id": {"$exists": False}}, {"$set": {"user_id": spv["user_id"]}}
        )
        await db.todo_entries.update_many(
            {"user_id": {"$exists": False}}, {"$set": {"user_id": spv["user_id"]}}
        )

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
